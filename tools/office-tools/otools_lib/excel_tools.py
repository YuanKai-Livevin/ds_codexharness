# -*- coding: utf-8 -*-
"""Excel 工具：合并 / 去重 / 筛选 / 透视 / 公式检查。"""

import os

import openpyxl

from .common import ToolError, cell_value, ensure_dir, list_paths, out, require, resolve, validate_output


def _load_rows(path, sheet=None):
    """读取工作表为行列表（首行为表头）。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
    except KeyError:
        wb.close()
        raise ToolError("SCHEMA_ERROR", f"工作表不存在: {sheet}（现有: {wb.sheetnames}）")
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell_value(c) for c in row])
    wb.close()
    if not rows:
        raise ToolError("EMPTY_INPUT", f"文件无数据行: {path}")
    return rows


def _write_rows(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([h if h is not None else "" for h in headers])
    for r in rows:
        ws.append(r)
    wb.save(path)
    wb.close()


def excel_merge(params, root, dry_run=False):
    p = require(["inputs", "output"], params, ["mode", "sheet", "header_rows"])
    inputs = list_paths(root, p["inputs"], exts={".xlsx", ".xls", ".csv"})
    if not inputs:
        raise ToolError("EMPTY_INPUT", "没有可合并的文件")
    out_path = resolve(root, p["output"])
    if os.path.splitext(out_path)[1].lower() not in (".xlsx", ".xls"):
        raise ToolError("SCHEMA_ERROR", "合并输出必须是 .xlsx 文件")
    header_rows = int(p.get("header_rows") or 1)
    if header_rows < 1:
        header_rows = 1
    plans = []
    for f in inputs:
        rows = _load_rows(f, p.get("sheet"))
        plans.append((os.path.basename(f), len(rows) - 1))
    if dry_run:
        return out(True, f"预演：合并 {len(inputs)} 个文件 → {out_path}",
                   outputs=[out_path], extra={"plans": plans})
    all_rows, headers = [], None
    for f in inputs:
        rows = _load_rows(f, p.get("sheet"))
        if headers is None:
            headers = rows[0][:header_rows][0] if header_rows == 1 else rows[0]
        data = rows[header_rows:]
        # 列数对齐
        if headers and data:
            data = [r[:len(headers)] + [""] * max(0, len(headers) - len(r)) for r in data]
        all_rows.extend(data)
    if not all_rows:
        raise ToolError("EMPTY_INPUT", "合并后无数据行")
    ensure_dir(os.path.dirname(out_path) or root)
    _write_rows(out_path, headers or [], all_rows)
    validate_output(out_path, "xlsx")
    return out(True, f"合并完成：{len(inputs)} 个文件共 {len(all_rows)} 行 → {out_path}",
               outputs=[out_path], extra={"files": len(inputs), "rows": len(all_rows)})


def excel_dedupe(params, root, dry_run=False):
    p = require(["input", "output"], params, ["key_columns"])
    src = resolve(root, p["input"])
    if not os.path.isfile(src):
        raise ToolError("FILE_NOT_FOUND", f"输入文件不存在: {p['input']}")
    out_path = resolve(root, p["output"])
    rows = _load_rows(src)
    headers = rows[0]
    data = rows[1:]
    keys = p.get("key_columns")
    if keys:
        idx = []
        for k in keys:
            if isinstance(k, int) or (isinstance(k, str) and k.isdigit()):
                i = int(k) - 1
            elif k in headers:
                i = headers.index(k)
            else:
                raise ToolError("SCHEMA_ERROR", f"key_columns 列不存在: {k}")
            if i < 0 or i >= len(headers):
                raise ToolError("SCHEMA_ERROR", f"列序号越界: {k}")
            idx.append(i)
    else:
        idx = list(range(len(headers)))
    seen, kept, dropped = set(), [], 0
    for r in data:
        key = tuple(str(r[i]) if i < len(r) else "" for i in idx)
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        kept.append(r)
    if dry_run:
        return out(True, f"预演：去重将保留 {len(kept)} 行，删除 {dropped} 行重复",
                   extra={"rows_before": len(data), "rows_after": len(kept), "dropped": dropped})
    ensure_dir(os.path.dirname(out_path) or root)
    _write_rows(out_path, headers, kept)
    validate_output(out_path, "xlsx")
    return out(True, f"去重完成：{len(data)} → {len(kept)} 行（删除 {dropped} 行）→ {out_path}",
               outputs=[out_path], extra={"rows_before": len(data), "rows_after": len(kept), "dropped": dropped})


def excel_filter(params, root, dry_run=False):
    p = require(["input", "output", "column", "op", "value"], params)
    src = resolve(root, p["input"])
    if not os.path.isfile(src):
        raise ToolError("FILE_NOT_FOUND", f"输入文件不存在: {p['input']}")
    out_path = resolve(root, p["output"])
    op = p["op"]
    if op not in ("eq", "ne", "gt", "ge", "lt", "le", "contains"):
        raise ToolError("SCHEMA_ERROR", f"op 必须是 eq|ne|gt|ge|lt|le|contains，当前: {op}")
    rows = _load_rows(src)
    headers = rows[0]
    col = p["column"]
    if isinstance(col, int) or (isinstance(col, str) and col.isdigit()):
        ci = int(col) - 1
    elif col in headers:
        ci = headers.index(col)
    else:
        raise ToolError("SCHEMA_ERROR", f"列不存在: {col}")
    if ci < 0 or ci >= len(headers):
        raise ToolError("SCHEMA_ERROR", f"列序号越界: {col}")
    want = p["value"]
    try:
        want_num = float(want)
    except (TypeError, ValueError):
        want_num = None
    kept = []
    for r in rows[1:]:
        v = r[ci] if ci < len(r) else ""
        hit = False
        try:
            vn = float(v) if v not in ("", None) else None
        except (TypeError, ValueError):
            vn = None
        if op == "eq":
            hit = (str(v) == str(want)) or (vn is not None and want_num is not None and vn == want_num)
        elif op == "ne":
            hit = not ((str(v) == str(want)) or (vn is not None and want_num is not None and vn == want_num))
        elif op in ("gt", "ge", "lt", "le"):
            if vn is None or want_num is None:
                hit = False
            elif op == "gt":
                hit = vn > want_num
            elif op == "ge":
                hit = vn >= want_num
            elif op == "lt":
                hit = vn < want_num
            else:
                hit = vn <= want_num
        elif op == "contains":
            hit = str(want).lower() in str(v).lower()
        if hit:
            kept.append(r)
    if dry_run:
        return out(True, f"预演：筛选后 {len(kept)}/{len(rows) - 1} 行",
                   extra={"rows_before": len(rows) - 1, "rows_after": len(kept)})
    ensure_dir(os.path.dirname(out_path) or root)
    _write_rows(out_path, headers, kept)
    validate_output(out_path, "xlsx")
    return out(True, f"筛选完成：{len(rows) - 1} → {len(kept)} 行（{col} {op} {want}）→ {out_path}",
               outputs=[out_path], extra={"rows_before": len(rows) - 1, "rows_after": len(kept)})


def excel_pivot(params, root, dry_run=False):
    p = require(["input", "output", "rows", "values"], params, ["agg"])
    src = resolve(root, p["input"])
    if not os.path.isfile(src):
        raise ToolError("FILE_NOT_FOUND", f"输入文件不存在: {p['input']}")
    out_path = resolve(root, p["output"])
    agg = p.get("agg") or "sum"
    if agg not in ("sum", "count", "avg"):
        raise ToolError("SCHEMA_ERROR", f"agg 必须是 sum|count|avg，当前: {agg}")
    rows = _load_rows(src)
    headers = rows[0]
    row_cols = p["rows"]
    if not isinstance(row_cols, list) or not row_cols:
        raise ToolError("SCHEMA_ERROR", "rows 必须是非空数组（分组列）")
    idx = []
    for c in list(row_cols) + [p["values"]]:
        if isinstance(c, int) or (isinstance(c, str) and c.isdigit()):
            i = int(c) - 1
        elif c in headers:
            i = headers.index(c)
        else:
            raise ToolError("SCHEMA_ERROR", f"列不存在: {c}")
        if i < 0 or i >= len(headers):
            raise ToolError("SCHEMA_ERROR", f"列序号越界: {c}")
        idx.append(i)
    ridx, vidx = idx[:-1], idx[-1]
    groups = {}
    for r in rows[1:]:
        key = tuple(str(r[i]) if i < len(r) else "" for i in ridx)
        try:
            v = float(r[vidx]) if r[vidx] not in ("", None) else 0.0
        except (TypeError, ValueError):
            v = 0.0
        g = groups.setdefault(key, [0, 0.0])
        g[1] += v
        g[0] += 1
    out_rows = []
    for key, (cnt, total) in sorted(groups.items(), key=lambda kv: kv[0]):
        if agg == "sum":
            val = total
        elif agg == "count":
            val = cnt
        else:
            val = total / cnt if cnt else 0
        out_rows.append(list(key) + [val])
    head = [headers[i] for i in ridx] + [f"{agg}:" + p["values"]]
    if dry_run:
        return out(True, f"预演：透视将产生 {len(out_rows)} 个分组",
                   extra={"groups": len(out_rows)})
    ensure_dir(os.path.dirname(out_path) or root)
    _write_rows(out_path, head, out_rows)
    validate_output(out_path, "xlsx")
    return out(True, f"透视完成：{len(out_rows)} 个分组（{agg} {p['values']}）→ {out_path}",
               outputs=[out_path], extra={"groups": len(out_rows)})


def excel_formula_check(params, root, dry_run=False):
    p = require(["input"], params)
    src = resolve(root, p["input"])
    if not os.path.isfile(src):
        raise ToolError("FILE_NOT_FOUND", f"输入文件不存在: {p['input']}")
    if dry_run:
        return out(True, f"预演：将检查公式错误 {src}")
    wb = openpyxl.load_workbook(src, data_only=False)
    errors = []
    formula_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str):
                    if v.startswith("="):
                        formula_count += 1
                    if v.startswith("#") and v.endswith(("!", "0!", "NAME?", "VALUE!", "NULL!",
                                                         "NUM!", "N/A", "REF!", "DIV/0!")):
                        errors.append({"sheet": ws.title, "cell": c.coordinate, "value": v})
    wb.close()
    if not errors:
        return out(True, f"公式检查通过：共 {formula_count} 个公式，无错误值",
                   extra={"formula_count": formula_count, "error_count": 0})
    return out(False, f"发现 {len(errors)} 个公式错误单元格（共 {formula_count} 个公式）",
               extra={"formula_count": formula_count, "error_count": len(errors), "errors": errors[:50]})
